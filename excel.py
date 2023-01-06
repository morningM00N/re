from typing import List
import openpyxl  
from tkinter import filedialog
from tkinter import messagebox
from openpyxl.utils.cell import range_boundaries
from tkinter.simpledialog import *
import datetime

def DP(boolean):
    if boolean==False:
        print("error")
        input()


intToChar = [None]
for i in range(26):
    intToChar.append(chr(ord('A')+i))

def getCellName(row,col):
    return intToChar[col]+str(row)

class Node:
    def __init__(self,name) -> None:
        self.name = name
        self.children = []
        self.colNames = []
        
    def drawAtSh(self,sh,row,col):
        if len(self.children)==0:
            sh[getCellName(row,col)].value=self.name
            return 1
        
        itCol = col
        for i in range(len(self.children)):
            ret = self.children[i].drawAtSh(sh,row+1,itCol)
            itCol+=ret
        if self.name == 'root':
            return
        for c in range(col, itCol):
            sh[getCellName(row,c)].value=self.name
        return itCol-col
            
        
    def clearColNames(self):
        self.colNames = []
        for i in range(len(self.children)):
            self.children[i].clearColNames()

    def computeColNames(self):
        if len(self.children)==0:
            self.colNames=[self.name]
        for i in range(len(self.children)):
            self.children[i].computeColNames()
            for j in range(len(self.children[i].colNames)):
                self.colNames.append(self.name+"_."+self.children[i].colNames[j])
            
        
    def insert(self, nameList: List[str], idx:int):
        if self.name != nameList[idx]:
            return False
        
        if len(nameList) <= idx+1:
            return True
        
        for i in range(len(self.children)):
            if self.children[i].insert(nameList,idx+1) == True:
                return True
        
        newNode = Node(nameList[idx+1])
        self.children.append(newNode)
        newNode.insert(nameList,idx+1)
        
        return True
    
    def merge(self, otherNode):
        DP(self.name == otherNode.name)
        lastMatch = 0
        for i in range(len(otherNode.children)):
            found = False
            for j in range(len(self.children)):
                if self.children[j].name == otherNode.children[i].name:
                    lastMatch = j
                    self.children[j].merge(otherNode.children[i])
                    found = True
                    break
            if found == False:
                self.children.insert(lastMatch,otherNode.children[i])
            
                
        
        

class Header:
    def __init__(self) -> None:
        self.root = Node('root')
        
    def insert(self, str: List[str]):
        """ col정보 (str)을 입력

        Args:
            str (List[str]): col 이름이 순서대로 적혀있는 List
            ex) ['주문중개','주문금액','바로결제주문금액']
        """
        self.root.insert(str,0)
    
    def merge(self, otherHeader):
        self.root.merge(otherHeader.root)
        pass
    
        
        

class RefinedData:
    def __init__(self) -> None:
        self.code = None
        self.header = Header()
        self.startRow = None
        self.startCol = None
        self.maxCol = None
        self.maxRow = None
        self.startDataRow = None
        self.data = []
        self.sh = None
        self.colName = None
        
    def readFile(self,filepath):
        self.code = filepath[filepath.rindex('/')+1:].split(' ')[0]
        wb = openpyxl.load_workbook(filepath)
        idx = 0
        while '상세' not in wb.sheetnames[idx]:
            idx+=1
        self.sh = wb[wb.sheetnames[idx]]

    
    def extractorHeader(self):
        ret = self.header
        sh = self.sh
        maxRow = sh.max_row
        maxCol = sh.max_column
        startRow = 1
        startCol = maxCol
        while sh[getCellName(startRow,maxCol-1)].value == None:
            startRow +=1

        while startCol>0 and sh[getCellName(maxRow,startCol)].value != None:
            startCol-=1
        startCol+= 1
        
        startDataRow = startRow
        while True:
            found = False
            for i in range(startCol, maxCol+1):
                if sh[getCellName(startDataRow,i)].value!=None and type(sh[getCellName(startDataRow,i)].value)!=str:
                    found = True
                    break
            if found == True:
                break
            startDataRow+=1
        
        prevTop = None    
        for col in range(startCol,maxCol+1):
            colHier=['root']
            for row in range(startRow, startDataRow):
                if sh[getCellName(row,col)].value==None:
                    for rng in sh.merged_cells.ranges:
                        if getCellName(row,col) in rng:
                            colHier.append(sh[rng.__str__().split(':')[0]].value)
                            break
                        # [minTemp,maxTemp] = rng.__str__().split(':')
                        # min_col = ord(minTemp[0])-ord('A')+1
                        # max_col = ord(maxTemp[0])-ord('A')+1
                        # min_row = int(minTemp[1:])
                        # max_row = int(maxTemp[1:])
                        # if 3>5:
                        #if col>=min_col and col <=max_col and row>=min_row and row<=max_row:
                            
                    
                else:
                    colHier.append(sh[getCellName(row,col)].value)
            prevTop = colHier[1]
            ret.insert(colHier)
            
        self.startRow = startRow
        self.startCol = startCol
        self.maxCol = maxCol
        self.maxRow = maxRow
        self.startDataRow = startDataRow
        self.colName = [0] * (self.maxCol-self.startCol +2)
        
        self.header.root.computeColNames()
        for col in range(self.startCol, self.maxCol+1):
            self.colName[col] = self.header.root.colNames[col-self.startCol].replace('root_.','')
        return

            
    def extractData(self):
        for row in range(self.startDataRow, self.maxRow+1):
            tempDict = {}
            for col in range(self.startCol, self.maxCol+1):
                tempDict[self.colName[col]] = self.sh[getCellName(row,col)].value
            self.data.append(tempDict)
        pass
        
if __name__ == "__main__":
    list_file = []
    files = filedialog.askopenfilenames(initialdir="./",\
                 title = "파일을 선택 해 주세요.",\
                    filetypes = (("*.xlsx","*xlsx"),("*.xls","*xls")))
    if files == '':
        messagebox.showwarning('경고','파일을 선택해주세요')
        exit()
    #files = ['KO012 황금점 배달의민족-2022년12월_정산명세서_박현종사장님.xlsx']
    
    baseHeader = None
    maxHeaderLen = 0
    errorFileList = []
    for f in files:
        try:
            exData = RefinedData()
            exData.readFile(f)
            exData.extractorHeader()
            exData.extractData()
            if maxHeaderLen < len(exData.colName):
                baseHeader = exData.header
                maxHeaderLen = len(exData.colName)
            list_file.append(exData)
        except:
            errorFileList.append(f)
    
    for ex in list_file:
        baseHeader.merge(ex.header)
    
    baseHeader.root.clearColNames()
    baseHeader.root.computeColNames()
    colNameMap = {}
    for i in range(len(baseHeader.root.colNames)):
        colNameMap[baseHeader.root.colNames[i].replace('root_.','')] = i+2

    newWb = openpyxl.Workbook()
    newSh = newWb.active
    # --- draw header
    baseHeader.root.drawAtSh(newSh,2,2)
    maxRow = newSh.max_row
    for r in range(3,maxRow+1):
        newSh[getCellName(r,1)] = '매장'
    
    for col in range(1,len(baseHeader.root.colNames)+2):
        for row in range(3,maxRow):
            startRow = row
            endRow = row+1
            while newSh[getCellName(startRow,col)].value == newSh[getCellName(endRow,col)].value:
                endRow+=1
            if startRow != endRow-1:
                temp = newSh[getCellName(startRow,col)].value
                newSh.merge_cells(start_row = startRow,start_column = col,end_row = endRow-1,end_column = col)
                newSh[getCellName(startRow,col)] = temp
            if endRow>=maxRow:
                break
    for row in range(3,maxRow+1):
        for col in range(1,len(baseHeader.root.colNames)+2):
            inRange = False
            for rng in newSh.merged_cells.ranges:
                if getCellName(row,col) in rng:
                    inRange=True
                    break
            if inRange:
                continue
            endCol = col+1
            while newSh[getCellName(row,col)].value == newSh[getCellName(row,endCol)].value:
                endCol+=1
            if col+1 != endCol:
                temp = newSh[getCellName(row,col)].value
                newSh.merge_cells(start_row = row,start_column = col,end_row = row,end_column = endCol-1)
                newSh[getCellName(row,col)] = temp
            if endCol>=len(baseHeader.root.colNames)+1:
                break
            

            
    
    startRow = maxRow+1
    for ex in list_file:
        for line in ex.data:
            newSh[getCellName(startRow,1)].value = ex.code
            for key in line:
                col = colNameMap[key]
                newSh[getCellName(startRow,col)].value = line[key]
            startRow+=1

    curTime = datetime.datetime.today().__str__().replace(':','')
    curTime = curTime[:curTime.index('.')]                
    
    outputfilename = askstring('파일명입력',"출력파일명을 입력하세요.")
    if outputfilename=="None" or outputfilename=="" or '.xlsx' not in outputfilename:
        
        outputfilename = "new_"+curTime+".xlsx"
    newWb.save(outputfilename)
    
    if len(errorFileList)!=0:
        messagebox.showwarning('에러','합치지 못한 파일이 있습니다. (실행 폴더를 확인해주세요.)')
        with open('합치지 못한 파일 '+curTime+'.txt', 'a+', -1, 'utf-8') as f:
            for i in range(len(errorFileList)):
                f.write(errorFileList[i]+"\n")
            f.close()


    messagebox.showwarning('종료','프로그램을 종료합니다.')
